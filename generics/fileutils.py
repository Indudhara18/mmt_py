import openpyxl
import json

'''
In order to handle the Excel file and Json file 
@author : Indudhara
@email  : indudhara18@gmail.com
'''

class FileUtils:

    # to create a new excel file
    def create_workbook(self,path):
        wb = openpyxl.Workbook()
        wb.save(path)

    # to get the name of excel sheet which is already created
    def get_sheetname(self,path):
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        return sheet.title

    # to read the data from existing excel sheet
    def readXLdata(self, path, sheetname, row, col):
        wb = openpyxl.load_workbook(path)
        sheet = wb.get_sheet_by_name(sheetname)
        data = sheet.cell(row=row, column=col).value
        return data

    # to write the dat into existing excel sheet
    def writeXLdata(self,path,sheetname,row,col,data):
        wb = openpyxl.load_workbook(path)
        sheet = wb.get_sheet_by_name(sheetname)
        sheet.cell(row=row,column=col).value=data
        wb.save(path)

    # to count no of data rows in excel work sheet
    def max_XLrows(self, path, sheetname):
        wb = openpyxl.load_workbook(path)
        sheet = wb.get_sheet_by_name(sheetname)
        row = sheet.max_row
        return row

    # to count no of data columns in excel work sheet
    def max_XLcol(self, path, sheetname):
        wb = openpyxl.load_workbook(path)
        sheet = wb.get_sheet_by_name(sheetname)
        column = sheet.max_column
        return column

    # to read the data from json file
    def readJson(self, jsonPath):
        with open(jsonPath, 'r') as file:
            data = json.load(file)
        return data

    # to write the data into json
    def writeJson(self, jsonPath, data):
        with open(jsonPath, 'w') as file:
            data = json.dump(data, file)
        return data


