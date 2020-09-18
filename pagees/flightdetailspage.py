import time
import openpyxl
from generics.webdriverutils import WebDriverUtils

'''
@author : Indudhara
email   : indudhara18@gmail.com
'''

class FlightPage:

    # xpath of webelements present in FlightDetailsPage
    __allflight = "//span[@class='airlineInfo-sctn']"
    __flightprices = "//span[@class='airlineInfo-sctn']/../../div[2]/div/div[3]"
    __passengerscls = "//input[@id='travellerAndClass']"
    __premiumeconomy = "//li[text()='Premium Economy']"
    __businessclass = "//li[text()='Business']"
    __firstclass = "//li[text()='First Class']"
    __done = "//button[text()='DONE']"
    __depflightname = "(//span[@class='font12 inlineB insertSep'])[1]"
    __depprice = "(//p[@class='actual-price'])[1]"
    __retflightname = "(//span[text()='Return'])[1]/../span[2]"
    __returnprice = "(//p[@class='actual-price'])[2]"
    __totalprice = "//span[@class='splitVw-total-fare']"
    #__booknow = "//button[@class='fli_primary_btn text-uppercase ' and text()='Book Now']"
    __booknow = "//div[@class='pull-left make_flex alC']"
    __contiue = "//button[text()='Continue']"


    def __init__(self,driver):
        self.driver = driver

    # utilization of webelements
    def __get_depflightname(self):
        return self.driver.find_element_by_xpath(self.__depflightname)

    def __get_depprice(self):
        return self.driver.find_element_by_xpath(self.__depprice)

    def __get_retflightname(self):
        return self.driver.find_element_by_xpath(self.__retflightname)

    def __get_retprice(self):
        return self.driver.find_element_by_xpath(self.__returnprice)

    def __get_totalprice(self):
        return self.driver.find_element_by_xpath(self.__totalprice)

    def get_booknow_button(self):
        return self.driver.find_element_by_xpath(self.__booknow)

    def get_continue_button(self):
        return self.driver.find_element_by_xpath(self.__contiue)

    def __get_passengercls(self):
        return self.driver.find_element_by_xpath(self.__passengerscls)

    def __get_done(self):
        return self.driver.find_element_by_xpath(self.__done)

    def __get_premiumeco(self):
        return self.driver.find_element_by_xpath(self.__premiumeconomy)

    def __get_businessclass(self):
        return self.driver.find_element_by_xpath(self.__businessclass)

    # to get the flight names and flight price
    def getflight_name_price(self):
        wdu = WebDriverUtils()
        wdu.get_all_elements_details(self.__allflight)
        wdu.get_all_elements_details(self.__flightprices)

    # to write flight name and flight price into excel
    def write_allflightdetails_intoexcel(self):
        path = r"C:\Users\indud\PycharmProjects\mmt\report\data.xlsx"
        book = openpyxl.load_workbook(path)
        sheet = book.get_sheet_by_name("Sheet1")

        sheet.cell(row=1,column=1).value="Flight Names"
        sheet.cell(row=1,column=2).value="Price"
        n = 2

        allfli = self.driver.find_elements_by_xpath(self.__allflight)
        for f in allfli:
            sheet.cell(row=n, column=1).value = f.text
            n += 1

        n = 2
        fliprices = self.driver.find_elements_by_xpath(self.__flightprices)
        for p in fliprices:
            sheet.cell(row=n, column=2).value = p.text
            n += 1
            book.save(path)

    # to traverse through seat types
    def seattype(self):
        self.__get_passengercls().click()
        self.__get_premiumeco().click()
        self.__get_done().click()
        time.sleep(1)
        self.__get_passengercls().click()
        self.__get_businessclass().click()
        self.__get_done().click()
        time.sleep(1)
        self.__get_passengercls().click()
        self.driver.find_element_by_xpath(self.__firstclass).click()
        self.__get_done().click()

    # to display flight details of both the sides in console
    def display_bothside_flightdetails(self):
        time.sleep(5)
        print("dep_flightname : ", self.__get_depflightname().text )
        print("dep_price : ", self.__get_depprice().text )
        print("ret_flightname : ", self.__get_retflightname().text)
        print("ret_price : ", self.__get_retprice().text)
        print("totalprice : ", self.__get_totalprice().text)

    # to write flight details of both the sides into excel
    def write_bothside_flightdetails(self):
        path = r"C:\Users\indud\PycharmProjects\mmt\report\data.xlsx"
        book = openpyxl.load_workbook(path)
        sheet = book.get_sheet_by_name("Sheet2")

        dep_flightname = self.__get_depflightname().text
        dep_price = self.__get_depprice().text
        ret_flightname = self.__get_retflightname().text
        ret_price = self.__get_retprice().text
        totalprice = self.__get_totalprice().text

        sheet.cell(row=1, column=1).value = "Departure"
        sheet.cell(row=2, column=1).value = dep_flightname
        sheet.cell(row=3, column=1).value = dep_price

        sheet.cell(row=1, column=2).value = "Return"
        sheet.cell(row=2, column=2).value = ret_flightname
        sheet.cell(row=3, column=2).value = ret_price

        sheet.cell(row=1, column=3).value = "Total Price"
        sheet.cell(row=3, column=3).value = totalprice
        book.save(path)

    # to get the total price
    def get_totalprice(self):
        time.sleep(2)
        return self.__get_totalprice().text